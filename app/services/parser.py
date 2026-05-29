import io
from openpyxl import load_workbook

REQUIRED_COLS = [
    "Employee ID", "Nickname / Name", "Cost Centre",
    "Gross Salary", "EPF Employer", "EIS Employer",
    "SOCSO Employer", "HRDF", "MTD", "CTC Hexa", "Net Salary",
]


def _to_num(val) -> float:
    if val is None or val == "":
        return 0.0
    try:
        return float(str(val).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def parse_excel_buffer(data: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    entities = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            continue

        header_row = rows[0]
        col_map: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell is not None:
                col_map[str(cell).strip()] = idx

        missing_cols = [c for c in REQUIRED_COLS if c not in col_map]

        employees = []
        for row in rows[1:]:
            if not row:
                continue
            emp_id_idx = col_map.get("Employee ID")
            if emp_id_idx is None:
                continue
            emp_id_val = row[emp_id_idx]
            if emp_id_val is None or str(emp_id_val).strip() == "":
                continue

            ctc_hexa = _to_num(row[col_map["CTC Hexa"]] if "CTC Hexa" in col_map else None)
            if ctc_hexa == 0:
                continue

            employees.append({
                "employeeId": str(emp_id_val).strip(),
                "name": str(row[col_map["Nickname / Name"]] or "").strip() if "Nickname / Name" in col_map else "",
                "costCentre": str(row[col_map["Cost Centre"]] or "").strip() if "Cost Centre" in col_map else "",
                "grossSalary": _to_num(row[col_map["Gross Salary"]] if "Gross Salary" in col_map else None),
                "epfEmployer": _to_num(row[col_map["EPF Employer"]] if "EPF Employer" in col_map else None),
                "eisEmployer": _to_num(row[col_map["EIS Employer"]] if "EIS Employer" in col_map else None),
                "socsoEmployer": _to_num(row[col_map["SOCSO Employer"]] if "SOCSO Employer" in col_map else None),
                "hrdf": _to_num(row[col_map["HRDF"]] if "HRDF" in col_map else None),
                "mtd": _to_num(row[col_map["MTD"]] if "MTD" in col_map else None),
                "ctcHexa": ctc_hexa,
                "netSalary": _to_num(row[col_map["Net Salary"]] if "Net Salary" in col_map else None),
            })

        if not employees:
            continue

        total_ctc = round(sum(e["ctcHexa"] for e in employees), 2)
        entities.append({
            "sheetName": sheet_name.strip(),
            "employees": employees,
            "totalCTC": total_ctc,
            "missingColumns": missing_cols,
        })

    wb.close()
    return entities
